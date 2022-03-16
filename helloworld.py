import math
import converter
from converter import kgs_to_lbs
from converter import kgs_to_lbs
from utils import find_max

from pathlib import Path

# print('mosh hamdani')
# print('o----')
# print(' ||||')
# print('*' * 10)
# price = 10
# rating = 10.2
# is_bilal = True
# print(price,rating)

# name = 'John Smith'
# age = 20
# status = 'New Patient'
# is_new = True

# name = input('Enter Your Name? ')
# print('hi' + name)
#
# age = input('Enter your year')
# actual_age = 2022 - int(age)
# print(type(age))
# print(actual_age)

# dakh = "bilal's laptop"
# na_dakh = 'this is the quotation from "bilal"'
# email = '''
# Hi john this is my first email in python
# bilalyounus33@gmail.com
# '''
# print(dakh + '  ' + na_dakh + '  ' + email)
# print(na_dakh[0])
# print(na_dakh[-1])
# print(na_dakh[0:3])

# formatable string
# first = 'Bilal'
# last = 'Younus'
# message  = first + '[' +  last +'] is a coder'
# print(message)
# msg = f'{first} [{last}] is a coder '
# print(msg)

# course = 'Python for Begineer'
# print(len(course))
# print(course.upper())
# print('P' in course
# print(10 ** 3)
# print(round(2.9))
# print(math.floor(2.3))

# is_boolean  = True
# if is_boolean:
#     print('this is true otherwise')
# elif:
#     print('this is an second if condition')
# else:
#     print('this is an else statament')
# print('this is false statement')

# has_good_income = 10000
# has_good_credit = True

# if has_good_income and has_good_credit
#     print('Eligilible for loan')
# if has_good_income and not has_good_credit
#     print('this is nit operator')
# converted/40 this gives floor
# converted//40 this gives int


# numbers = 1
# while numbers < 5:
#     print(numbers)
#     numbers +=1

# guess_count = 0
# guess_correct = 9
# while guess_count < 3 :
#     guess = int(input('Enter Yout Guess'))
#     guess_count+=1
#     if guess == guess_correct :
#         print('You hace right guess')
#         break
# else:
#     print('galat ha yha')

# command=''
# started = False
# stop = False
# while True:
#     command = input('Enter Command ').lower()
#     if command == 'start' and started == False:
#         print('the car has been started')
#         started = True
#     elif command == 'stop' and stop == False:
#        print('the car has been stoped')
#        stop=True;
#     elif command == 'quit':
#         break ;
#     elif command == 'help':
#         print("""
# start -- to start car
# stop -- to stop the car
# help -- to see the option
# """
# )
#     else :
#         print('Sorry didnt understand')
#

# print('P' in course)
# for item in 'Python':
#     print(item)
# for item2 in ['bilal','fahad','yamin']:
#     print(item2)
# for item3 in range(10):
#     print(item3)
# for items in [10,20,30]:
#     items+=items
# print(items)
#

# for x in range(3):
#     for y in range(2):
#         print(f"( {x}, {y} )")

# now we will print the f
from typing import List

value = [5, 2, 5, 2, 2]

# for i in value:
#     i=str(i)
#     for j in i:
#         print('*' * int(i))
#
# for i in value:
#     output= '*'
#     for j in range(i):
#         output+='*'
#     print(output)


# list
# list = ['Bilal', 'Younus', 'Rana', 'Rao', 'Raj put']
# print(list)
# print(list[0])
# print(list[2:])
# list[0] = 'Ahsan'
# print(list)

# find the largest number in the list
# li = [1, 2, 3, 4, 5, 6, 7, 8, 9]
# save = li[0]
# for i in li:
#     if save < i:
#         save = i
#
# print(save)

# 2d list

# number = [
#     [1, 2, 3],
#     [4, 5, 6],
#     [7 ,8, 9]
# ]
# print(number[0][2])
#
# numbers = [1, 2, 3, 4, 5, 6]
# numbers.append(12)
# numbers.insert(0,13)
# print(numbers)

# check the duplicate item in the list

numbersss = [2, 6, 3, 5, 7, 6, 3, 7]
count = 0;
match = ''

# for i in numbersss:
#     match= i
#     for j in numbersss:
#         if match == j :
#             count+=1
#             if count >=2:
#                 numbersss.remove(j)
#     count =0
# print(numbersss)

# numbers = []
# for i in numbersss:
#     if i not in numbers:
#         numbers.append(i)
# print(numbers)

# tuple
# numbers = (1, 2, 3)
# print(numbers[0])

# unpacking
# coordinate = (1, 2, 3)
# x,y,z=coordinate
# print(f'value of {x} , {y} , {z}')
# coordiantes = [1, 2, 3]
# s,d,f = coordiantes
# print(f'value of {s} , {d} , {f}')

# dictionary in python

# customer = {
#     'name': 'Bilal',
#     'age': 20 ,
#     'is_verified': True
# }
# print(customer['name'])
# print(customer.get('name'))
# customer['birthday'] = '22 jan'
# print(customer['birthday'])
# dict = {
#     '1': 'One',
#     '2': 'Two',
#     '3': 'Three' ,
#     '4': 'Four'
# }
# inp = input('Phone')
# output = ''
# for ch in inp:
#     output+=dict.get(ch,'!')
# print(output)

# emoji converter
# message = input('>')
# words = message.split(' ')
# dictionary = {
#     ':)' : '🤣' ,
#     ':(' : '😒'
# }
# output = ''
# for i in words:
#     output+= dictionary.get(i, i)
# print(output)

# functions
# def name():
#     print('This is in the function')
#     print('Do you know me')
#
#
# print('start')
# name()
# print('stop')

# parameters
# def name(name):
#     print(f'This is in the function {name}')
#     print('Do you know me')
#
#
# print('start')
# name('Bilal')
# print('stop')
#

# keyword arguments
#
# def name(first_name, last_name):
#     print('This is in the function')
#     print('Do you know me')
#     print(f'my name is {first_name}   {last_name}')
#
#
# print('start')
# name(first_name='Bilal', last_name='Younus')
# print('stop')

# return statement
# def square(number):
#     return  number * number
# name = square(3)
# print(name)

# creating reusable function

# exception
# try:
#     age = int(input('Enter Your Age'))
#     print(age)
# except ValueError:
#     print('Invalid Value')
# except ZeroDivisionError:
#     print('number can nto be zero')

# classess and constructor
# class Point:
#
#     def __init__(self,l,m):
#         self.l = l
#         self.m = m
#     def move(self):
#         print('move')
#     def draw(self):
#         print('draw')
#
#
# point1 =  Point(12,12)
# point1.move()
# point1.draw()
# point1.x = 20
# point1.y = 30
# print(point1.x , point1.y)
#
# #cinsturctor
# point2 = Point(30,40)
# print(point2.l , point2.m)

# class Person:
#     def __init__(self,name):
#         print('Constructor')
#         self.name = name
#     def talk(self):
#         print(f'This is {self.name}')
#
#
# person = Person('Bilal')
# person.talk()
#

# inheritance
# class Mammal:
#     def talk(self):
#         print('talk')
#
# class Dog(Mammal):
#     def bark(self):
#         print('brak')
#
#
# class Cat(Mammal):
#     def cry(self):
#         print('dont hesitate')
#
# dog1 = Dog()
# dog1.talk()
# cat = Cat()
# cat.cry()

# modules
# def lbs_to_kg(weight):
#     return weight * 0.45
#
#
# def kgs_to_lbs(weight):
#     return weight/0.45

# just have to import it
# this is the second method


# print(converter.lbs_to_kg(70))
# print(converter.kgs_to_lbs(150))
# because of the second import we dnt have to write the name of the module


# print(kgs_to_lbs(100))

# challenge
# find_max([3, 4, 7, 2, 6, 7])


# packages
# package is a container for multiple modules
# first well create an directory
# __init__.py in that file
# or you can directly select an option from the right click menu

# now you can follow two rules to import that things
# import ecommerce.shipping
# ecommerce.shipping.ship()

# from ecommerce.shipping import ship
#
# ship()
#
# from ecommerce import shipping
# shipping.ship()

# generating random variable

# python module index
import random
#
# for i in range(3):
#     print(random.random())
#     print(random.randint(10, 20))
# num = ['Fahad', 'Yamin', 'Ahsan','Bilal', 'Abdullah']
# leader = random.choice(num)
# print(leader )

#challenge

#
# class Dice:
#     def roll(self):
#         store = (int(random.random() * 6 + 1),int(random.random() * 6 + 1) )
#         print(store)
#
# dice = Dice()
# dice.roll()

#files and directory
#absolute path
#C:/bilal/programs/
#relative path
#
#import pathlib from Path
# path = Path('ecommerce')
# print(path.exists())

# path = Path('emails')
# print(path.rmdir())
#
# path = Path()
# print(path.glob('*.*'))
# for file in path.glob('*.*'):
#     print(file)
#
#

#pypi and pip


#excel spreadsheet

import openpyxl as excel

from openpyxl.chart import BarChart, reference, Reference

wb = excel.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1, 1)
print(cell.value)
print(sheet.max_row)
for row in range(2, sheet.max_row +1):
    # print(row)
    cell = sheet.cell(row,3)
    corrected = cell.value * 0.90
    print(corrected)
    corrected_price_cell = sheet.cell(row,4)
    corrected_price_cell.value = corrected
    print(corrected_price_cell)

values =Reference(sheet , min_row=2 , max_row=sheet.max_row, min_col=4 ,max_col=4)
chart = BarChart()
chart.add_data(values)
chart.add_chart(chart, 'e2')

wb.save('transaction2.xlsx')

